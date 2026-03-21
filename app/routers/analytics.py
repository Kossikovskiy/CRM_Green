# GrassCRM — app/routers/analytics.py v8.0.1

import io
import os
from collections import defaultdict
from datetime import date
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import extract, func
from sqlalchemy.orm import Session as DBSession, joinedload

from app.database import get_db
from app.models import (
    Deal, DealService, DealComment, Stage, Contact,
    Expense, ExpenseCategory, Budget,
)
from app.schemas import BudgetCreate, BudgetUpdate, BudgetResponse
from app.security import require_admin, guard_project, is_won_stage, is_lost_stage

router = APIRouter()

MONTHS_RU = ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"]


# ── ВСПОМОГАТЕЛЬНАЯ ───────────────────────────────────────────────────────────
def _calc_time_stats(won_deals: list) -> dict:
    dh = [d for d in won_deals if d.duration_hours and d.duration_hours > 0]
    total_h = sum(d.duration_hours for d in dh)
    total_r = sum((d.revenue if d.revenue else d.total) or 0 for d in dh)
    return {
        "total_hours":        round(total_h, 1),
        "deals_with_hours":   len(dh),
        "revenue_per_hour":   round(total_r / total_h, 0) if total_h else None,
        "avg_hours_per_deal": round(total_h / len(dh), 1) if dh else None,
    }


# ── АНАЛИТИКА ─────────────────────────────────────────────────────────────────
@router.get("/api/analytics/funnel")  # kept for backwards compat
def get_funnel(year: int, project: str = "pokos", db: DBSession = Depends(get_db), user: dict = Depends(require_admin)):
    guard_project(project, user)
    return get_analytics(year, project, db, user)


@router.get("/api/analytics")
def get_analytics(year: int, project: str = "pokos", db: DBSession = Depends(get_db), user: dict = Depends(require_admin)):
    guard_project(project, user)

    stages = db.query(Stage).filter(Stage.project == project).order_by(Stage.order).all()
    deals  = (db.query(Deal)
              .options(joinedload(Deal.services).joinedload(DealService.service),
                       joinedload(Deal.materials), joinedload(Deal.contact))
              .filter(extract("year", Deal.deal_date) == year, Deal.project == project)
              .all())

    won_stage_ids  = {s.id for s in stages if is_won_stage(s)}
    lost_stage_ids = {s.id for s in stages if is_lost_stage(s)}
    won  = [d for d in deals if d.stage_id in won_stage_ids]
    lost = [d for d in deals if d.stage_id in lost_stage_ids]

    total_revenue = sum((d.revenue if d.revenue else d.total) or 0 for d in won)
    avg_check     = round(total_revenue / len(won), 2) if won else 0
    win_rate      = round(len(won) / len(deals) * 100, 1) if deals else 0

    # Воронка
    funnel = [{"stage_id": s.id, "stage_name": s.name, "color": s.color, "is_final": s.is_final,
               "count": sum(1 for d in deals if d.stage_id == s.id),
               "total": sum(d.total or 0 for d in deals if d.stage_id == s.id)}
              for s in stages]

    # Помесячная динамика
    rev_by_month: dict = defaultdict(float)
    for d in won:
        dt = d.deal_date or d.closed_at or d.created_at
        if dt: rev_by_month[dt.month] += (d.revenue if d.revenue else d.total) or 0

    expenses_year = db.query(Expense).filter(extract("year", Expense.date) == year, Expense.project == project).all()
    exp_by_month: dict = defaultdict(float)
    for e in expenses_year: exp_by_month[e.date.month] += e.amount

    monthly = [{"month": i, "label": MONTHS_RU[i-1],
                "revenue": round(rev_by_month[i], 2), "expenses": round(exp_by_month[i], 2),
                "profit": round(rev_by_month[i] - exp_by_month[i], 2)}
               for i in range(1, 13)]

    # Топ услуг
    svc_rev: dict = defaultdict(float)
    svc_cnt: dict = defaultdict(int)
    svc_names: dict = {}
    for d in won:
        for ds in d.services:
            sname = ds.service.name if ds.service else f"#{ds.service_id}"
            svc_rev[ds.service_id] += ds.price_at_moment * ds.quantity
            svc_cnt[ds.service_id] += 1
            svc_names[ds.service_id] = sname
    top_services = sorted(
        [{"id": sid, "name": svc_names[sid], "revenue": round(svc_rev[sid], 2), "count": svc_cnt[sid]} for sid in svc_rev],
        key=lambda x: x["revenue"], reverse=True
    )[:8]

    # Расходы по категориям
    cats = {c.id: c.name for c in db.query(ExpenseCategory).all()}
    exp_by_cat: dict = defaultdict(float)
    for e in expenses_year: exp_by_cat[cats.get(e.category_id, "Прочее")] += e.amount
    expense_by_category = sorted([{"name": k, "amount": round(v, 2)} for k, v in exp_by_cat.items()], key=lambda x: x["amount"], reverse=True)

    # Повторные клиенты
    repeat_won = [d for d in won if d.is_repeat]
    new_won    = [d for d in won if not d.is_repeat]

    # Причины провала
    lost_reason_count: dict = defaultdict(int)
    if lost_stage_ids:
        for (txt,) in (db.query(DealComment.text).join(Deal, Deal.id == DealComment.deal_id)
                       .filter(extract("year", Deal.deal_date) == year,
                               Deal.stage_id.in_(lost_stage_ids),
                               DealComment.text.ilike("Причина провала:%")).all()):
            reason = (txt or "").split(":", 1)[1].strip() if ":" in (txt or "") else (txt or "Не указана")
            lost_reason_count[reason or "Не указана"] += 1

    # Источники клиентов
    client_source_count: dict = defaultdict(int)
    unique_cids = {d.contact_id for d in deals if d.contact_id}
    if unique_cids:
        for c in db.query(Contact).filter(Contact.id.in_(list(unique_cids))).all():
            client_source_count[(c.source or "").strip() or "Не указан"] += 1

    return {
        "total_deals": len(deals), "won_deals": len(won), "lost_deals": len(lost),
        "win_rate": win_rate, "avg_check": avg_check,
        "total_revenue": round(total_revenue, 2),
        "total_expenses": round(sum(e.amount for e in expenses_year), 2),
        "funnel": funnel, "monthly": monthly,
        "top_services": top_services, "expense_by_category": expense_by_category,
        "repeat": {
            "repeat_count": len(repeat_won), "new_count": len(new_won),
            "repeat_revenue": round(sum((d.revenue if d.revenue else d.total) or 0 for d in repeat_won), 2),
            "new_revenue": round(sum((d.revenue if d.revenue else d.total) or 0 for d in new_won), 2),
            "repeat_rate": round(len(repeat_won) / len(won) * 100, 1) if won else 0,
        },
        "lost_reasons": sorted([{"name": k, "count": v} for k, v in lost_reason_count.items()], key=lambda x: x["count"], reverse=True),
        "client_sources": sorted([{"name": k, "count": v} for k, v in client_source_count.items()], key=lambda x: x["count"], reverse=True),
        "time": _calc_time_stats(won),
    }


@router.get("/api/analytics/clients")
def get_client_margin(year: int, project: str = "pokos", db: DBSession = Depends(get_db), user: dict = Depends(require_admin)):
    guard_project(project, user)
    stages = db.query(Stage).filter(Stage.project == project).all()
    won_ids = {s.id for s in stages if is_won_stage(s)}
    deals = (db.query(Deal).options(joinedload(Deal.contact), joinedload(Deal.materials))
             .filter(Deal.project == project, Deal.stage_id.in_(won_ids),
                     Deal.is_archived == False, extract("year", Deal.deal_date) == year).all())

    stats: dict = defaultdict(lambda: {"name": "", "deals": 0, "revenue": 0.0, "hours": 0.0, "mat_cost": 0.0})
    for d in deals:
        cid = d.contact_id or 0
        stats[cid]["name"]     = d.contact.name if d.contact else "Без клиента"
        stats[cid]["deals"]   += 1
        stats[cid]["revenue"] += (d.revenue if d.revenue else d.total) or 0
        stats[cid]["hours"]   += d.duration_hours or 0
        stats[cid]["mat_cost"] += sum((m.cost_price or 0) * (m.quantity or 0) for m in d.materials)

    result = [{"contact_id": cid, "name": s["name"], "deals": s["deals"],
               "revenue": round(s["revenue"], 2), "mat_cost": round(s["mat_cost"], 2),
               "margin": round(s["revenue"] - s["mat_cost"], 2), "hours": round(s["hours"], 1),
               "revenue_per_hour": round(s["revenue"] / s["hours"], 0) if s["hours"] else None}
              for cid, s in stats.items()]
    return {"clients": sorted(result, key=lambda x: x["revenue"], reverse=True)[:20]}


@router.get("/api/analytics/weekdays")
def get_weekday_load(year: int, project: str = "pokos", db: DBSession = Depends(get_db), user: dict = Depends(require_admin)):
    guard_project(project, user)
    DAYS_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    stages = db.query(Stage).filter(Stage.project == project).all()
    won_ids = {s.id for s in stages if is_won_stage(s)}
    deals = (db.query(Deal).filter(Deal.project == project, Deal.stage_id.in_(won_ids),
                                   Deal.is_archived == False, extract("year", Deal.deal_date) == year,
                                   Deal.deal_date.isnot(None)).all())
    counts: dict = defaultdict(int)
    revenue: dict = defaultdict(float)
    hours: dict = defaultdict(float)
    for d in deals:
        wd = d.deal_date.weekday()
        counts[wd] += 1
        revenue[wd] += (d.revenue if d.revenue else d.total) or 0
        hours[wd]   += d.duration_hours or 0
    return {"days": [{"day": i, "label": DAYS_RU[i], "deals": counts[i],
                      "revenue": round(revenue[i], 0), "hours": round(hours[i], 1)} for i in range(7)]}


@router.get("/api/analytics/seasons")
def get_seasons(project: str = "pokos", db: DBSession = Depends(get_db), user: dict = Depends(require_admin)):
    guard_project(project, user)
    SEASON_MONTHS = list(range(4, 11))
    MONTHS_MAP = {4:"Апр", 5:"Май", 6:"Июн", 7:"Июл", 8:"Авг", 9:"Сен", 10:"Окт"}
    current_year = date.today().year
    years = [current_year - 2, current_year - 1, current_year]
    stages = db.query(Stage).filter(Stage.project == project).all()
    won_ids = {s.id for s in stages if is_won_stage(s)}
    deals = (db.query(Deal).filter(Deal.project == project, Deal.stage_id.in_(won_ids),
                                   Deal.is_archived == False,
                                   extract("year", Deal.deal_date).in_(years),
                                   extract("month", Deal.deal_date).in_(SEASON_MONTHS)).all())
    revenue: dict = {y: defaultdict(float) for y in years}
    hours: dict   = {y: defaultdict(float) for y in years}
    for d in deals:
        if not d.deal_date: continue
        y, m = d.deal_date.year, d.deal_date.month
        revenue[y][m] += (d.revenue if d.revenue else d.total) or 0
        if d.duration_hours: hours[y][m] += d.duration_hours
    return {
        "years": years,
        "months": [{"month": m, "label": MONTHS_MAP[m]} for m in SEASON_MONTHS],
        "revenue": {str(y): [round(revenue[y][m], 0) for m in SEASON_MONTHS] for y in years},
        "hours":   {str(y): [round(hours[y][m],   1) for m in SEASON_MONTHS] for y in years},
    }


# ── БЮДЖЕТ ────────────────────────────────────────────────────────────────────
@router.get("/api/budget", response_model=List[BudgetResponse])
def get_budget(year: int, project: str = "pokos", db: DBSession = Depends(get_db), user: dict = Depends(require_admin)):
    guard_project(project, user)
    return db.query(Budget).filter(Budget.year == year, Budget.project == project).order_by(Budget.id).all()


@router.post("/api/budget", status_code=201)
def create_budget(data: BudgetCreate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    item = Budget(**data.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return item


@router.patch("/api/budget/{item_id}")
def update_budget(item_id: int, data: BudgetUpdate, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    item = db.query(Budget).filter(Budget.id == item_id).first()
    if not item: raise HTTPException(404, "Запись не найдена")
    for k, v in data.model_dump(exclude_unset=True).items(): setattr(item, k, v)
    db.commit(); db.refresh(item)
    return item


@router.delete("/api/budget/{item_id}", status_code=204)
def delete_budget(item_id: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    item = db.query(Budget).filter(Budget.id == item_id).first()
    if item: db.delete(item); db.commit()
    return None


# ── ЭКСПОРТ EXCEL ─────────────────────────────────────────────────────────────
def _openpyxl_available():
    try:
        import openpyxl; return True
    except ImportError:
        return False


@router.get("/api/export/excel")
def export_excel(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    if not _openpyxl_available():
        raise HTTPException(503, "openpyxl не установлен на сервере")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    GREEN = "1a3318"; SAGE = "4a7c3f"; LIGHT = "f5f2eb"; WHITE = "FFFFFF"

    def make_header(ws, cols, hc=GREEN):
        for c, (title, width) in enumerate(cols, 1):
            cell = ws.cell(1, c, title)
            cell.font      = Font(bold=True, color=WHITE, size=10)
            cell.fill      = PatternFill("solid", fgColor=hc)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.row_dimensions[1].height = 28

    def style_row(ws, row, r, alt=False):
        for c in range(1, len(row)+1):
            cell = ws.cell(r, c)
            cell.fill      = PatternFill("solid", fgColor="F0ECE3" if alt else WHITE)
            cell.alignment = Alignment(vertical="center")
            thin = Side(style="thin", color="E2DDD4")
            cell.border = Border(bottom=Border(bottom=thin).bottom)

    stages   = {s.id: s for s in db.query(Stage).all()}
    contacts = {c.id: c for c in db.query(Contact).all()}
    deals_q  = db.query(Deal).filter(extract("year", Deal.deal_date) == year).order_by(Deal.deal_date).all()

    # Лист 1: Сделки
    ws1 = wb.create_sheet("Сделки")
    make_header(ws1, [("ID",6),("Название",30),("Клиент",22),("Этап",18),("Сумма",14),("Дата создания",18),("Дата закрытия",18),("Менеджер",18)])
    for i, d in enumerate(deals_q, 2):
        row = [d.id, d.title, contacts.get(d.contact_id, type("x",(),{"name":"—"})()).name,
               stages.get(d.stage_id, type("x",(),{"name":"—"})()).name, d.total or 0,
               d.created_at.strftime("%d.%m.%Y") if d.created_at else "",
               d.closed_at.strftime("%d.%m.%Y") if d.closed_at else "", d.manager or ""]
        for c, v in enumerate(row, 1): ws1.cell(i, c, v)
        style_row(ws1, row, i, i%2==0)
    ws1.freeze_panes = "A2"; ws1.auto_filter.ref = f"A1:H{len(deals_q)+1}"

    # Лист 2: Расходы
    ws2 = wb.create_sheet("Расходы")
    expenses_q = db.query(Expense).filter(extract("year", Expense.date) == year).order_by(Expense.date).all()
    cats = {c.id: c.name for c in db.query(ExpenseCategory).all()}
    make_header(ws2, [("ID",6),("Дата",14),("Название",35),("Категория",20),("Сумма",14)])
    for i, e in enumerate(expenses_q, 2):
        row = [e.id, e.date.strftime("%d.%m.%Y") if e.date else "", e.name, cats.get(e.category_id,"—"), e.amount]
        for c, v in enumerate(row, 1): ws2.cell(i, c, v)
        style_row(ws2, row, i, i%2==0)
    ws2.freeze_panes = "A2"
    tr = len(expenses_q)+2
    ws2.cell(tr, 4, "ИТОГО").font = Font(bold=True)
    ws2.cell(tr, 5, sum(e.amount for e in expenses_q)).font = Font(bold=True)

    # Лист 3: Аналитика
    ws3 = wb.create_sheet("Аналитика")
    won_ids = {s.id for s in stages.values() if is_won_stage(s)}
    won_deals = [d for d in deals_q if d.stage_id in won_ids]
    total_rev = sum(d.total or 0 for d in won_deals)
    total_exp = sum(e.amount for e in expenses_q)
    ws3.column_dimensions["A"].width = 35; ws3.column_dimensions["B"].width = 22
    ws3.cell(1,1,"СВОДКА ЗА ГОД").font = Font(bold=True, size=13, color=GREEN)
    ws3.row_dimensions[1].height = 26
    for r, (label, val) in enumerate([("Год", year),("Всего сделок",len(deals_q)),("Успешных",len(won_deals)),
                                       ("Выручка",total_rev),("Расходы",total_exp),("Прибыль",total_rev-total_exp),
                                       ("Средний чек",round(total_rev/len(won_deals),2) if won_deals else 0)], 2):
        ws3.cell(r,1,label).font = Font(bold=True, size=10)
        c2 = ws3.cell(r,2,val); c2.fill = PatternFill("solid", fgColor=LIGHT); c2.alignment = Alignment(horizontal="right")

    # Лист 4: Бюджет
    ws4 = wb.create_sheet("Бюджет")
    budgets_q = db.query(Budget).filter(Budget.year == year).all()
    make_header(ws4, [("Название",28),("Период",14),("Плановая выручка",20),("Плановые расходы",20),("Прибыль",20),("Заметки",30)])
    for i, b in enumerate(budgets_q, 2):
        row = [b.name, b.period, b.planned_revenue or 0, b.planned_expenses or 0,
               (b.planned_revenue or 0)-(b.planned_expenses or 0), b.notes or ""]
        for c, v in enumerate(row, 1): ws4.cell(i, c, v)
        style_row(ws4, row, i, i%2==0)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="grasscrm_{year}.xlsx"'})


# ── ЭКСПОРТ PDF ───────────────────────────────────────────────────────────────
@router.get("/api/export/pdf")
def export_pdf(year: int, db: DBSession = Depends(get_db), _=Depends(require_admin)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, KeepTogether
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        font_ok = False
        for fp in ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                   "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                   "/usr/share/fonts/TTF/DejaVuSans.ttf"]:
            if os.path.exists(fp):
                try: pdfmetrics.registerFont(TTFont("F", fp))
                except Exception: pass
                font_ok = True; break
        F = "F" if font_ok else "Helvetica"
    except ImportError:
        raise HTTPException(503, "reportlab не установлен на сервере")

    stages     = {s.id: s for s in db.query(Stage).all()}
    contacts   = {c.id: c for c in db.query(Contact).all()}
    deals_q    = db.query(Deal).options(joinedload(Deal.contact)).filter(extract("year", Deal.deal_date) == year).order_by(Deal.deal_date).all()
    expenses_q = db.query(Expense).options(joinedload(Expense.category)).filter(extract("year", Expense.date) == year).order_by(Expense.date).all()

    won_ids   = {s.id for s in stages.values() if is_won_stage(s)}
    lost_ids  = {s.id for s in stages.values() if is_lost_stage(s)}
    won_deals = [d for d in deals_q if d.stage_id in won_ids]
    lost_deals= [d for d in deals_q if d.stage_id in lost_ids]
    total_rev = sum(d.total or 0 for d in won_deals)
    total_exp = sum(e.amount or 0 for e in expenses_q)
    profit    = total_rev - total_exp
    avg_check = total_rev / len(won_deals) if won_deals else 0
    win_rate  = round(len(won_deals) / len(deals_q) * 100, 1) if deals_q else 0

    def rub(v): return f"{float(v or 0):,.0f}\u202f₽".replace(",", "\u202f")

    C_DARK=colors.HexColor("#1a3318"); C_GREEN=colors.HexColor("#3d6b35"); C_SAGE=colors.HexColor("#4a7c3f")
    C_LIGHT=colors.HexColor("#f5f2eb"); C_ALT=colors.HexColor("#ede9e0"); C_WHITE=colors.white
    C_MUTED=colors.HexColor("#8a8070"); C_BORDER=colors.HexColor("#d9d4c8")
    W = A4[0] - 4*cm

    def ps(name, **kw):
        d = dict(fontName=F, fontSize=9, leading=13, textColor=C_DARK); d.update(kw)
        return ParagraphStyle(name, **d)

    S_LOGO=ps("logo",fontSize=20,leading=24,spaceAfter=2); S_H2=ps("h2",fontSize=10,textColor=C_GREEN,spaceBefore=18,spaceAfter=5)
    S_BODY=ps("body",fontSize=8.5,leading=12); S_FOOT=ps("foot",fontSize=7,textColor=C_MUTED)
    S_KPI_L=ps("kpil",fontSize=7,textColor=C_MUTED,leading=10,spaceAfter=1)
    S_KPI_V=ps("kpiv",fontSize=14,leading=17,spaceAfter=0); S_KPI_V2=ps("kpiv2",fontSize=11,textColor=C_SAGE,leading=14,spaceAfter=0)

    def make_table(rows, widths, hdr=C_DARK, repeat=1):
        t = Table(rows, colWidths=widths, repeatRows=repeat)
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),hdr),("TEXTCOLOR",(0,0),(-1,0),C_WHITE),
            ("FONTNAME",(0,0),(-1,-1),F),("FONTSIZE",(0,0),(-1,0),7.5),("FONTSIZE",(0,1),(-1,-1),8.5),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("LINEBELOW",(0,0),(-1,0),0.8,C_SAGE),("LINEBELOW",(0,1),(-1,-1),0.3,C_BORDER),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[C_WHITE,C_ALT]),("BOX",(0,0),(-1,-1),0.5,C_BORDER),
        ])); return t

    story = []
    hdr_t = Table([[Paragraph("GrassCRM",S_LOGO), Paragraph(f"Отчёт за {year} год<br/><font color='#8a8070' size='8'>Сформирован {date.today().strftime('%d.%m.%Y')}</font>",S_BODY)]],colWidths=[W*0.5,W*0.5])
    hdr_t.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),F),("VALIGN",(0,0),(-1,-1),"BOTTOM"),("ALIGN",(1,0),(1,0),"RIGHT"),("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0)]))
    story.append(hdr_t); story.append(HRFlowable(width=W,thickness=1.5,color=C_SAGE,spaceAfter=10,spaceBefore=6))

    def kpi(label, val, sub=None):
        items = [Paragraph(label,S_KPI_L), Paragraph(val,S_KPI_V)]
        if sub: items.append(Paragraph(sub,S_KPI_V2))
        return items

    def kpi_table(data):
        rows = [[item[0] for item in data],[item[1] for item in data],[item[2] if len(item)>2 else Paragraph("",S_KPI_V2) for item in data]]
        t = Table(rows, colWidths=[W/4]*4)
        t.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),F),("BACKGROUND",(0,0),(-1,-1),C_LIGHT),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),("LEFTPADDING",(0,0),(-1,-1),12),("RIGHTPADDING",(0,0),(-1,-1),12),("LINEAFTER",(0,0),(-2,-1),0.5,C_BORDER),("BOX",(0,0),(-1,-1),0.5,C_BORDER)]))
        return t

    story.append(kpi_table([kpi("ВЫРУЧКА",rub(total_rev)),kpi("РАСХОДЫ",rub(total_exp)),kpi("ПРИБЫЛЬ",rub(profit)),kpi("СРЕДНИЙ ЧЕК",rub(avg_check))]))
    story.append(Spacer(1,4))
    story.append(kpi_table([kpi("СДЕЛОК ВСЕГО",str(len(deals_q))),kpi("КОНВЕРСИЯ",f"{win_rate}%",f"успешных {len(won_deals)}"),kpi("ПРОВАЛЕНО",str(len(lost_deals))),kpi("УНИ­КАЛ. КЛИЕНТЫ",str(len({d.contact_id for d in deals_q if d.contact_id})))]))
    story.append(Spacer(1,2))

    story.append(Paragraph("Помесячная динамика",S_H2)); story.append(HRFlowable(width=W,thickness=0.5,color=C_BORDER,spaceAfter=4))
    mrows=[["МЕСЯЦ","ВЫРУЧКА","РАСХОДЫ","ПРИБЫЛЬ","МАРЖА"]]
    for i,mn in enumerate(MONTHS_RU,1):
        mr=sum((d.total or 0) for d in won_deals if (d.deal_date or d.closed_at or d.created_at) and (d.deal_date or d.closed_at or d.created_at).month==i)
        me=sum((e.amount or 0) for e in expenses_q if e.date and e.date.month==i)
        mg=f"{round((mr-me)/mr*100,1)}%" if mr else "—"
        mrows.append([mn,rub(mr),rub(me),rub(mr-me),mg])
    story.append(KeepTogether(make_table(mrows,[2.5*cm,4.1*cm,4.1*cm,4.1*cm,3.2*cm]))); story.append(Spacer(1,4))

    story.append(Paragraph("Воронка продаж",S_H2)); story.append(HRFlowable(width=W,thickness=0.5,color=C_BORDER,spaceAfter=4))
    total_cnt=len(deals_q) or 1
    frows=[["ЭТАП","СДЕЛОК","СУММА","ДОЛЯ"]]
    for st in sorted(stages.values(),key=lambda s:s.order):
        cnt=sum(1 for d in deals_q if d.stage_id==st.id); amt=sum(d.total or 0 for d in deals_q if d.stage_id==st.id)
        frows.append([st.name,str(cnt),rub(amt),f"{round(cnt/total_cnt*100,1)}%"])
    story.append(KeepTogether(make_table(frows,[8*cm,2.8*cm,5*cm,2.2*cm],C_SAGE))); story.append(Spacer(1,4))

    story.append(Paragraph("Топ-10 успешных сделок",S_H2)); story.append(HRFlowable(width=W,thickness=0.5,color=C_BORDER,spaceAfter=4))
    drows=[["НАЗВАНИЕ","КЛИЕНТ","СУММА","ЗАКРЫТА"]]
    for d in sorted(won_deals,key=lambda x:x.total or 0,reverse=True)[:10]:
        cname=contacts.get(d.contact_id,type("_",(),{"name":"—"})()).name
        drows.append([d.title[:38],cname[:22],rub(d.total or 0),d.closed_at.strftime("%d.%m.%y") if d.closed_at else "—"])
    if len(drows)==1: drows.append(["Нет данных","—","—","—"])
    story.append(KeepTogether(make_table(drows,[8*cm,4.5*cm,3.5*cm,2*cm]))); story.append(Spacer(1,4))

    story.append(Paragraph("Расходы по категориям",S_H2)); story.append(HRFlowable(width=W,thickness=0.5,color=C_BORDER,spaceAfter=4))
    exp_by_cat: dict = {}
    for e in expenses_q:
        cat = e.category.name if e.category else "Без категории"
        exp_by_cat[cat] = exp_by_cat.get(cat, 0) + (e.amount or 0)
    te = total_exp or 1
    crows=[["КАТЕГОРИЯ","СУММА","ДОЛЯ"]]
    for cat,amt in sorted(exp_by_cat.items(),key=lambda x:x[1],reverse=True):
        crows.append([cat,rub(amt),f"{round(amt/te*100,1)}%"])
    if len(crows)==1: crows.append(["Нет данных",rub(0),"0%"])
    story.append(KeepTogether(make_table(crows,[9.5*cm,5*cm,3.5*cm],C_SAGE)))

    story.append(Spacer(1,20)); story.append(HRFlowable(width=W,thickness=0.5,color=C_BORDER,spaceAfter=4))
    story.append(Paragraph(f"GrassCRM · crmpokos.ru · Отчёт {year} года",S_FOOT))

    buf = io.BytesIO()
    SimpleDocTemplate(buf,pagesize=A4,leftMargin=2*cm,rightMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm).build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="grasscrm_{year}.pdf"',
                                      "Cache-Control": "no-store", "Pragma": "no-cache"})
